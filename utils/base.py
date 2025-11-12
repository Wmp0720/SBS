import json
import openpyxl, os

wb = openpyxl.Workbook()
sheet1 = wb.create_sheet("分析报告")
sheet1["A1"] = "问题"
sheet1["B1"] = "category"
sheet1["C1"] = "答案"
sheet1["D1"] = "modelName"
sheet1["E1"] = "sessionId"
sheet1["F1"] = "得分点"
sheet1["G1"] = "参考答案"
sheet1["H1"] = "score"
sheet1["I1"] = "score_reason"
sheet1["J1"] = "auto_score"
sheet1["K1"] = "auto_score_reason"
sheet1["L1"] = "分差"


def txt_to_excel(file):
    with open(file, "r", encoding="utf-8") as f:
        i = 2
        for line in f:
            _ = json.loads(line)
            sheet1["A{}".format(i)] = _["问题"]
            sheet1["B{}".format(i)] = _["category"]
            sheet1["C{}".format(i)] = _["答案"]
            sheet1["D{}".format(i)] = _["模型"]
            sheet1["E{}".format(i)] = _["sessionid"]
            if "得分点" in _.keys():
                sheet1["F{}".format(i)] = _["得分点"]
            if "参考答案" in _.keys():
                sheet1["G{}".format(i)] = _["参考答案"]
            sheet1["H{}".format(i)] = _["评分"]
            if "评价" in _.keys():
                sheet1["I{}".format(i)] = _["评价"]
            if "自动化score" in _.keys():
                sheet1["J{}".format(i)] = _["自动化score"]
            if "自动化score_reason" in _.keys():
                sheet1["K{}".format(i)] = _["自动化score_reason"]
            if "分数分差" in _.keys():
                sheet1["L{}".format(i)] = _["分数分差"]
            i += 1
    wb.save(file.replace("txt", "xlsx"))


def existExcel(fileName, project=None):
    if os.path.exists(fileName):
        workbook = openpyxl.load_workbook(fileName)
        sheet = workbook["分析报告"]
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("分析报告")
        if project == "端侧语言模型":
            header = ["模型", "sessionid", "session_sort", "category", "问题", "参考答案", "答案", "得分点", "评分", "评价",
                      "自动化score", "自动化score_reason", "分数分差", "prompt"]
        elif project == "多模态":
            header = ["模型", "sessionid", "sessionsort", "category", "问题", "参考答案", "答案", "得分点", "评分", "评价", "自动化score",
                      "自动化score_reason", "分数分差", "prompt", "prompt_length"]
        else:
            header = ["模型", "sessionid", "category", "问题", "参考答案", "答案", "得分点", "评分", "评价", "自动化score",
                      "自动化score_reason", "分数分差", "prompt"]
        for col in range(1, len(header) + 1):
            sheet.cell(row=1, column=col, value=header[col - 1])
        workbook.save(fileName)
    return workbook, sheet


def existFile(fileName):
    if os.path.isfile(fileName):
        return True
    else:
        return False


def getResult(data, project=None):
    sort_values = []
    if project == "多模态":
        sort_key = ["模型", "sessionid", "sessionsort", "category", "问题", "参考答案", "答案", "得分点", "评分", "评价", "自动化score",
                    "自动化score_reason", "分数分差", "prompt", "prompt_length"]
    elif project == "端侧语言模型":
        sort_key = ["model", "session_id", "session_sort", "分组维度", "query", "参考答案", "answer", "得分点", "分数均分", "人工评价",
                    "自动化score", "自动化score_reason", "分数分差", "评分prompt"]
    else:
        sort_key = ["模型", "sessionid", "category", "问题", "参考答案", "答案", "得分点", "评分", "评价", "自动化score", "自动化score_reason",
                    "分数分差", "prompt"]
    for key in sort_key:
        if key in data.keys():
            sort_values.append(data.get(key))
        elif key == "人工评价":
            judge = "标注员A评价：{}\n标注员B评价:{}".format(data.get("标注员A评价"), data.get("标注员B评价"))
            sort_values.append(judge)
        else:
            sort_values.append('')
    return sort_values