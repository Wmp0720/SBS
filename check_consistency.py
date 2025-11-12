
import os
import re
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl.styles import Font, Border, Side, PatternFill


# =========================================================
# 1. 公共工具
# =========================================================
def _safe_sheet_name(name: str) -> str:
    """生成合法且 ≤31 字符的 Sheet 名"""
    name = re.sub(r'[\[\]\:\*\?\/\\]', '_', str(name))
    return name[:31] if len(name) > 31 else name


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    将输入文件的列名统一为脚本内部使用的标准列名。
    """
    col_map = {
        # 维度列
        '度量三级维度': '度量三级分类',
        '度量四级维度': '度量四级分类',
        # 自研模型 ("小v") 相关列
        '小v满意度': '标注员_小v满意度',
        '小v优质弱智': '标注员_小v优质弱智',
        '标注员_小V优质弱智': '标注员_小v优质弱智',
        '小v优质弱智主要问题': '标注员_小v优质弱智主要问题',
        '标注员_小V优质弱智主要问题': '标注员_小v优质弱智主要问题',
        '小v多轮': '标注员_小v多轮',
        '小V链接': '小v链接',  # 处理大小写
        '小v主要问题': '标注员_小v主要问题',
        '小v主要问题一级分类': '标注员_小v主要问题一级分类',
        '标注员_小V主要问题一级分类': '标注员_小v主要问题一级分类',
        '小v竞品对比': '标注员_小v竞品对比',
        # 竞品模型相关列
        '竞品满意度': '标注员_竞品满意度',
        '竞品优质弱智': '标注员_竞品优质弱智',
        '竞品优质弱智主要问题': '标注员_竞品优质弱智主要问题',
        '竞品多轮': '标注员_竞品多轮',
        '竞品主要问题': '标注员_竞品主要问题',
        '竞品主要问题一级分类': '标注员_竞品主要问题一级分类',
        '竞品对比': '标注员_竞品竞品对比',
        # 其他
        '标注员标注结果': '标注结果'
    }

    # 使用rename函数，它会自动忽略字典中不存在于DataFrame列里的键
    df = df.rename(columns=col_map)
    print("已根据规则自动重命名列...")
    return df

def _calc_precision_recall(y_true, y_pred, labels):
    """返回 {label: {'recall': r, 'precision': p}}（均为浮点数, 0~1）"""
    out = {}
    for lab in labels:
        tp = ((y_true == lab) & (y_pred == lab)).sum()
        fn = ((y_true == lab) & (y_pred != lab)).sum()
        fp = ((y_true != lab) & (y_pred == lab)).sum()
        recall = tp / (tp + fn) if (tp + fn) else 0.0
        precision = tp / (tp + fp) if (tp + fp) else 0.0
        out[lab] = {'recall': recall, 'precision': precision}
    return out


def _calculate_primary_label_jaccard(df: pd.DataFrame, col_true: str, col_pred: str) -> float:
    """
    使用Jaccard相似度计算主要问题标签的一致率。
    - 规则1: 在计算前，所有标签回归到一级。
    - 规则2 (特殊): 如果一方只标了'9'，另一方只标了'13'，视为完全一致 (1.0)。
    """
    if df.empty:
        return 0.0

    def get_primary_set(label_string: str) -> set:
        if pd.isna(label_string) or str(label_string).strip() == '':
            return set()
        parts = re.split(r'[,，]', str(label_string))
        primary_labels = {part.strip().split('_', 1)[0] for part in parts if part.strip()}
        return primary_labels

    def jaccard_similarity(row):
        """
        计算主要问题标签的一致性。行级评判，返回1.0（一致）或0.0（不一致）。
        """
        set1 = get_primary_set(row[col_true])
        set2 = get_primary_set(row[col_pred])

        # 特殊情况1: 处理 '9' 和 '13' 的兼容
        if len(set1) == 1 and len(set2) == 1 and set1.union(set2) == {'9', '13'}:
            return 1.0

        # 特殊情况2: 双方都认为没问题
        if not set1 and not set2:
            return 1.0

        if set1.intersection(set2):
            return 1.0
        else:
            return 0.0


    jaccard_scores = df.apply(jaccard_similarity, axis=1)
    return jaccard_scores.mean()

# =========================================================
# 核心统计 —— 生成 7 张表（所有数值均为『数字』而非格式化字符串）
# =========================================================
def _generate_reports(df_src: pd.DataFrame,
                      model_name: str,
                      group_col_for_table2: str):
    total = len(df_src)

    # 准备对齐列
    sv_satis = (pd.to_numeric(df_src["标注员_小v满意度"], errors='coerce')
                == pd.to_numeric(df_src["LLMs_自研满意度"], errors='coerce')).sum()
    cp_satis = (pd.to_numeric(df_src["标注员_竞品满意度"], errors='coerce')
                == pd.to_numeric(df_src["LLMs_竞品满意度"], errors='coerce')).sum()

    # 满意率对比回归到一级标签
    sv_match = (df_src["标注员_小v优质弱智"].astype(str).str.split('_', n=1).str[0]
                == df_src["LLMs_自研优质弱智"].astype(str).str.split('_',n= 1).str[0]).sum()
    cp_match = (df_src["标注员_竞品优质弱智"].astype(str).str.split('_', n=1).str[0]
                == df_src["LLMs_竞品优质弱智"].astype(str).str.split('_', n=1).str[0]).sum()

    # 在满意率对比中加入 9 和 13 的特殊处理
    def calculate_match_with_exception(df, col_true, col_pred):
        p1_true = df[col_true].astype(str).str.split('_', n=1).str[0]
        p1_pred = df[col_pred].astype(str).str.split('_', n=1).str[0]

        # 条件1: 一般情况，一级标签相等
        general_match = (p1_true == p1_pred)

        # 条件2: 特殊情况，一个为'9'，另一个为'13'
        special_match_9_13 = ((p1_true == '9') & (p1_pred == '13')) | \
                             ((p1_true == '13') & (p1_pred == '9'))

        # 两者满足其一即为一致
        final_match = general_match | special_match_9_13
        return final_match.sum()

    sv_match = calculate_match_with_exception(df_src, "标注员_小v优质弱智", "LLMs_自研优质弱智")
    cp_match = calculate_match_with_exception(df_src, "标注员_竞品优质弱智", "LLMs_竞品优质弱智")

    # 胜负平对比保持原样，因为它们本身就是一级标签
    sv_vs = (df_src["标注员_小v竞品对比"].astype(str).str.strip()
             == df_src["LLMs_自研竞品对比"].astype(str).str.strip()).sum()
    cp_vs = (df_src["标注员_竞品竞品对比"].astype(str).str.strip()
             == df_src["LLMs_竞品竞品对比"].astype(str).str.strip()).sum()

    # ---------- 表1 总体人机一致率 ----------
    tbl1 = pd.DataFrame([
        {"指标": "合格率（0-1）", "一致率": (sv_satis + cp_satis) / (2 * total) if total else 0},
        {"指标": "满意率（弱智、不合格、合格、优质）", "一致率": (sv_match + cp_match) / (2 * total) if total else 0},
        {"指标": "胜出率", "一致率": (sv_vs + cp_vs) / (2 * total) if total else 0},
    ])

    # ---------- 表2 分垂类一致率 ----------
    rows2 = []
    rows2.append({
        "维度": "总计",
        "样本数": total,
        "合格率": (sv_satis + cp_satis) / (2 * total) if total else 0,
        "满意率": (sv_match + cp_match) / (2 * total) if total else 0,
        "胜出率": (sv_vs + cp_vs) / (2 * total) if total else 0
    })
    for dim, g in df_src.groupby(group_col_for_table2):
        cnt = len(g)
        if cnt == 0: continue
        sv_s = (pd.to_numeric(g["标注员_小v满意度"], errors='coerce')
                == pd.to_numeric(g["LLMs_自研满意度"], errors='coerce')).sum()
        cp_s = (pd.to_numeric(g["标注员_竞品满意度"], errors='coerce')
                == pd.to_numeric(g["LLMs_竞品满意度"], errors='coerce')).sum()

        # [修改点 1] ========= START =========
        sv_m = (g["标注员_小v优质弱智"].astype(str).str.split('_', n=1).str[0]
                == g["LLMs_自研优质弱智"].astype(str).str.split('_',n=1).str[0]).sum()
        cp_m = (g["标注员_竞品优质弱智"].astype(str).str.split('_', n=1).str[0]
                == g["LLMs_竞品优质弱智"].astype(str).str.split('_', n=1).str[0]).sum()
        # [修改点 1] ========= END =========

        comp_m = (g["标注员_竞品竞品对比"].astype(str).str.strip()
                  == g["LLMs_竞品竞品对比"].astype(str).str.strip()).sum()
        rows2.append({
            "维度": dim,
            "样本数": cnt,
            "合格率": (sv_s + cp_s) / (2 * cnt),
            "满意率": (sv_m + cp_m) / (2 * cnt),
            "胜出率": comp_m / cnt
        })
    tbl2 = pd.DataFrame(rows2)
# ---------- 表3 分竞品一致率 ----------
    tbl3 = pd.DataFrame([
        {"竞品": "蓝心小v",
         "合格率": sv_satis / total if total else 0,
         "满意率": sv_match / total if total else 0,
         "胜出率": sv_vs / total if total else 0},
        {"竞品": "豆包",
         "合格率": cp_satis / total if total else 0,
         "满意率": cp_match / total if total else 0,
         "胜出率": cp_vs / total if total else 0},
    ])

    # ---------- 表4 & 5 召回率精确率 -----------
    def _tbl_01(df, col_true, col_pred, prefix):
        # ... (此函数无需修改)
        rows = []
        y_t = pd.to_numeric(df[col_true], errors='coerce')
        y_p = pd.to_numeric(df[col_pred], errors='coerce')
        m = _calc_precision_recall(y_t, y_p, [0, 1])
        rows.append({"维度": "总计", "样本数": len(df),
                     f"{prefix}_0召回率": m[0]['recall'],
                     f"{prefix}_0精确率": m[0]['precision'],
                     f"{prefix}_1召回率": m[1]['recall'],
                     f"{prefix}_1精确率": m[1]['precision']})
        for dim, g in df.groupby(group_col_for_table2):
            y_t = pd.to_numeric(g[col_true], errors='coerce')
            y_p = pd.to_numeric(g[col_pred], errors='coerce')
            m = _calc_precision_recall(y_t, y_p, [0, 1])
            rows.append({"维度": dim, "样本数": len(g),
                         f"{prefix}_0召回率": m[0]['recall'],
                         f"{prefix}_0精确率": m[0]['precision'],
                         f"{prefix}_1召回率": m[1]['recall'],
                         f"{prefix}_1精确率": m[1]['precision']})
        return pd.DataFrame(rows)

    tbl4 = _tbl_01(df_src, "标注员_小v满意度", "LLMs_自研满意度", "自研")
    tbl5 = _tbl_01(df_src, "标注员_竞品满意度", "LLMs_竞品满意度", "竞品")

    # ---------- 表6 胜负平召回率 ----------
    y_true_c = df_src["标注员_竞品竞品对比"].astype(str).str.strip()
    y_pred_c = df_src["LLMs_竞品竞品对比"].astype(str).str.strip()
    m_c = _calc_precision_recall(y_true_c, y_pred_c, ["胜", "负", "平"])
    tbl6_rows = []
    for lab in ["胜", "负", "平"]:
        consistency = ((y_true_c == lab) & (y_pred_c == lab)).sum() / (y_true_c == lab).sum() if (
                                                                                                             y_true_c == lab).sum() > 0 else 0
        tbl6_rows.append({
            "维度": lab,
            "原样本数": (y_true_c == lab).sum(),
            "人机一致率": consistency,
            "自研_召回率": m_c[lab]['recall'],
            "自研_精确率": m_c[lab]['precision'],
        })
    tbl6 = pd.DataFrame(tbl6_rows)

    # ---------- 表7 主要问题一致率 ----------
    tbl7 = None
    required_cols = {'标注员_小v主要问题', 'LLMs_自研主要问题', '标注员_竞品主要问题', 'LLMs_竞品主要问题'}
    if required_cols.issubset(df_src.columns):
        rows7 = []

        # 使用新的 Jaccard 函数进行计算
        sv_questions_jaccard = _calculate_primary_label_jaccard(df_src, "标注员_小v主要问题", "LLMs_自研主要问题")
        cp_questions_jaccard = _calculate_primary_label_jaccard(df_src, "标注员_竞品主要问题", "LLMs_竞品主要问题")

        rows7.append({
            "维度": "总计",
            "原样本数": total,
            "主要问题人机一致率": (sv_questions_jaccard + cp_questions_jaccard) / 2 if total else 0
        })
        for dim, g in df_src.groupby(group_col_for_table2):
            sv_q_j = _calculate_primary_label_jaccard(g, "标注员_小v主要问题", "LLMs_自研主要问题")
            cp_q_j = _calculate_primary_label_jaccard(g, "标注员_竞品主要问题", "LLMs_竞品主要问题")
            rows7.append({
                "维度": dim,
                "原样本数": len(g),
                "主要问题人机一致率": (sv_q_j + cp_q_j) / 2 if len(g) > 0 else 0
            })

        tbl7 = pd.DataFrame(rows7)

    return tbl1, tbl2, tbl3, tbl4, tbl5, tbl6, tbl7, total


# =========================================================
# 样式
# =========================================================
font_cn = Font(name='微软雅黑', size=11)
thin_side = Side(style='thin', color="99CCFF")
border_all = Border(left=thin_side, right=thin_side,
                    top=thin_side, bottom=thin_side)
header_fill = PatternFill("solid", fgColor="D9D9D9")
firstrow_fill = PatternFill("solid", fgColor="BDD7EE")


def _style_range(ws, min_row, max_row, min_col, max_col, df, first_data_row):
    for r in ws.iter_rows(min_row=min_row, max_row=max_row,
                          min_col=min_col, max_col=max_col):
        for cell in r:
            cell.font = font_cn
            cell.border = border_all
            # 检查列是否存在于DataFrame中
            if (cell.column - 1) < len(df.columns):
                col_name = df.columns[cell.column - 1]
                if cell.row == min_row:
                    cell.fill = header_fill
                elif cell.row == first_data_row:
                    cell.fill = firstrow_fill

                if col_name in ("样本数", "总数(未包含剔除数据)", "原样本数"):
                    cell.number_format = '0'
                else:
                    cell.number_format = '0.00%'


def _write_table(writer, sheet_name, ws, start_row, title, df):
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value=f"{title}（无数据）").font = font_cn
        return start_row + 2

    ws.cell(row=start_row, column=1, value=title).font = Font(name='微软雅黑', size=11, bold=True)
    df.to_excel(writer, sheet_name=sheet_name,
                index=False, startrow=start_row + 1)

    header_row = start_row + 2
    first_data_row = header_row + 1

    max_row = header_row + len(df)

    max_col = df.shape[1]

    _style_range(ws, header_row, max_row, 1, max_col, df, first_data_row)

    return max_row + 3


def _write_one_sheet(writer, sheet_name, model_name,
                     tbl1, tbl2, tbl3, tbl4, tbl5, tbl6, tbl7, total):
    if sheet_name in writer.book.sheetnames:
        del writer.book[sheet_name]
    ws = writer.book.create_sheet(sheet_name)
    writer.sheets[sheet_name] = ws

    cur = 1
    if sheet_name == "总人机一致率":
        ws.cell(1, 1, "蓝心小vSBS自动化评分人机一致率").font = Font(name="微软雅黑", bold=True, size=18)
        ws.cell(2, 1, f"评分时间：{datetime.now().strftime('%Y年%m月%d日')}").font = Font(name="微软雅黑", bold=True,
                                                                                         size=12)
        ws.cell(3, 1, f"评分模型：{model_name}").font = Font(name="微软雅黑", bold=True, size=12)
        ws.cell(4, 1, f"数据集总量：{total}").font = Font(name="微软雅黑", bold=True, size=12)
        cur = 6
    cur = _write_table(writer, sheet_name, ws, cur, "表1：总体人机一致率", tbl1)
    cur = _write_table(writer, sheet_name, ws, cur, "表2：分垂类一致率", tbl2)
    cur = _write_table(writer, sheet_name, ws, cur, "表3：分竞品一致率", tbl3)
    cur = _write_table(writer, sheet_name, ws, cur, "表4：蓝心小v 0/1 召回率与精确率", tbl4)
    cur = _write_table(writer, sheet_name, ws, cur, "表5：豆包 0/1 召回率与精确率", tbl5)
    cur = _write_table(writer, sheet_name, ws, cur, "表6：胜负平标签 召回率与精确率", tbl6)
    if tbl7 is not None:
        _write_table(writer, sheet_name, ws, cur, "表7：主要问题人机一致率", tbl7)


def add_consistency_flag_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    为DataFrame添加三列独立的'人机一致'标记列（优化版）。

    “人机一致_满意度评级”列的逻辑已更新，以支持更宽松的匹配规则：
    - 参考 _calculate_primary_label_jaccard 的逻辑。
    - 特殊处理：'9'和'13'视为一致。
    """
    print("正在添加三列'人机一致'标记 (优化版)...")

    # --- 1. 规范化所有待比较的列（保持不变） ---
    human_sbs = df['标注员_小v竞品对比'].astype(str).str.strip()
    llm_sbs = df['LLMs_自研竞品对比'].astype(str).str.strip()
    human_v_rating = df['标注员_小v优质弱智'].astype(str).str.strip()
    llm_v_rating = df['LLMs_自研优质弱智'].astype(str).str.strip()
    human_c_rating = df['标注员_竞品优质弱智'].astype(str).str.strip()
    llm_c_rating = df['LLMs_竞品优质弱智'].astype(str).str.strip()
    human_v_binary = df['标注员_小v满意度'].astype(str).str.strip()
    llm_v_binary = df['LLMs_自研满意度'].astype(str).str.strip()
    human_c_binary = df['标注员_竞品满意度'].astype(str).str.strip()
    llm_c_binary = df['LLMs_竞品满意度'].astype(str).str.strip()

 # --- 2. 为“满意度评级”定义一个更宽松的匹配函数 ---
    def lenient_rating_match(human_series, llm_series) -> pd.Series:
        """
        接收两个Pandas Series，返回一个布尔型的Series，表示每行是否匹配。
        """
        # 将每个单元格的字符串按逗号分割成标签集合
        human_sets = human_series.str.split('[,，]').apply(lambda x: set(s.strip() for s in x if s.strip()))
        llm_sets = llm_series.str.split('[,，]').apply(lambda x: set(s.strip() for s in x if s.strip()))

        # 将两个集合Series合并，以便逐行比较
        comparison_df = pd.concat([human_sets.rename('human'), llm_sets.rename('llm')], axis=1)

        def row_wise_check(row):
            h_set, l_set = row['human'], row['llm']

            # 规则1：处理 '9' 和 '13' 的特殊兼容情况
            if (h_set == {'9'} and l_set == {'13'}) or \
                    (h_set == {'13'} and l_set == {'9'}):
                return True

            # 规则2：双方都为空集，视为一致（例如，都判定为“13无问题”或均为空）
            if not h_set and not l_set:
                return True

            # 规则3
            if h_set.intersection(l_set):
                return True

            return False

        # 对每一行应用我们的宽松匹配规则
        return comparison_df.apply(row_wise_check, axis=1)

    # --- 3. 定义各维度的一致性条件（应用新函数） ---

    # 条件1: 胜负平一致（逻辑不变）
    sbs_match = (human_sbs == llm_sbs)

    # 条件2: 满意度评级一致 (*** 使用新的宽松匹配逻辑 ***)
    v_rating_match = lenient_rating_match(human_v_rating, llm_v_rating)
    c_rating_match = lenient_rating_match(human_c_rating, llm_c_rating)
    rating_match = v_rating_match & c_rating_match  # 最终评级一致性需要自研和竞品都满足

    # 条件3: 合格率(0/1)一致（逻辑不变）
    binary_match = (human_v_binary == llm_v_binary) & (human_c_binary == llm_c_binary)

    # --- 4. 使用 np.where 高效地生成新列（保持不变） ---
    df['人机一致_胜负平'] = np.where(sbs_match, '一致', '不一致')
    df['人机一致_满意度评级'] = np.where(rating_match, '一致', '不一致')
    df['人机一致_合格率'] = np.where(binary_match, '一致', '不一致')

    print("三列'人机一致'标记列添加完成！")
    return df


# =========================================================
# 入口
# =========================================================
def compute_consistency(file_path: str, model_name: str) -> None:
    df_raw = pd.read_excel(file_path, sheet_name=0)
    df_raw = _normalize_columns(df_raw)
    # 剔除“剔除”标签
    df_raw = df_raw[~df_raw['标注员_小v满意度'].astype(str).isin(['剔除'])
                    & ~df_raw['LLMs_自研满意度'].astype(str).isin(['剔除'])].copy()
    if df_raw.empty:
        print("无有效数据")
        return
    with pd.ExcelWriter(file_path,
                        engine="openpyxl",
                        mode='a' if os.path.exists(file_path) else 'w',
                        if_sheet_exists='overlay') as writer:
        res = _generate_reports(df_raw, model_name, '度量一级分类')
        _write_one_sheet(writer, "总人机一致率", model_name, *res)
        for p_dim, df_grp in df_raw.groupby("度量一级分类"):
            sheet_name = _safe_sheet_name(p_dim)
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]
            res_sub = _generate_reports(df_grp, model_name, '度量二级分类')
            _write_one_sheet(writer, sheet_name, model_name, *res_sub)
    print("人机一致率相关 Sheet 已全部写入完成！")


if __name__ == "__main__":
    file_path = r"D:\My Documents\11179872\Desktop\度量\代码\automated-evaluation\Results\先验知识_初始版\test1_o3\test1_o3Eval.xlsx"
    compute_consistency(file_path, model_name="o3")